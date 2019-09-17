#envia_respuestas_wa.py 

from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as EC

import time
import openpyxl as excel

driver = webdriver.Chrome('chromedriver.exe') 
driver.get('https://web.whatsapp.com/')
input("apreta algun boton luego de escanear el QR")

  

def busca_elemetos ():
  
  text = ''
  textfinder1 = driver.find_elements_by_class_name('_3La1s')
  for i in textfinder1:
    text += i.text + '\n' 
  print(text)


def envia_mensaje (contact, text):
  inp_xpath_search = "//input[@title='Buscar o empezar un chat nuevo']"
  input_box_search = WebDriverWait(driver,50).until(lambda driver: driver.find_element_by_xpath(inp_xpath_search))
  input_box_search.click()
  time.sleep(2)
  input_box_search.send_keys(contact)
  time.sleep(2)
  try:
      cont_a_sel = driver.find_element_by_class_name('//div[@class="X7YrQ" style="z-indez: 0"]')
      
      tit = cont_a_sel.text
      pass
  except NoSuchElementException:
      print("Oops! no esta ese contacto...")
      bot_clear = '//button[@class="_2heX1"]'
      time.sleep(2)
      boton_clear = driver.find_element_by_xpath(bot_clear)
      time.sleep(2)
      boton_clear.click()
      ws['D' + str(i)] = "no enviado"
      file.save("contacts.xlsx")
      return

  while True:
    try:
      selected_contact = driver.find_element_by_xpath("//span[@title='"+tit+"']")
      selected_contact.click()
      inp_xpath = '//div[@class="_3u328 copyable-text selectable-text"][@contenteditable="true"][@data-tab="1"]'
      input_box = driver.find_element_by_xpath(inp_xpath)
      time.sleep(2)
      # input_box.send_keys(text + Keys.ENTER)
      input_box.send_keys(text)
      bot_clear = '//button[@class="_2heX1"]'
      time.sleep(2)
      boton_clear = driver.find_element_by_xpath(bot_clear)
      time.sleep(2)
      boton_clear.click()
      ws['D' + str(i)] = "enviado"
      file.save("contacts.xlsx")
      time.sleep(2)
      return
    except:
      ws['D' + str(i)] = "no enviado"
      file.save("contacts.xlsx")
      pass

file = excel.load_workbook(filename = "contacts.xlsx")
ws = file.active
i=0
for row in ws.values:
  contacto = row[0]
  texto = row[1]
  print(row[0]+" con el mensaje "+row[1])
  i=i+1
  envia_mensaje (contacto,texto)  






