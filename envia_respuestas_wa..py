from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl as excel

def menu_principal(opcion):
  print("_"*40)
  print("1. Filtrar contactos para envio")
  print("2. Enviar mensajes via enlace")
  print("3. Enviar mensaje a contactos\n")
  while True:
    try:
      opcion = int(input("Selecciona una opcion:"))
      if opcion < 4 and opcion > 0:
        return(opcion)
      else:
        print("Opcion incorrecta 1-3")
        menu_principal(opcion=0)
        pass
    except ValueError:
        print("Opcion incorrecta 1-3")
  
def busca_elemetos():

  text = ''
  textfinder1 = driver.find_elements_by_class_name('_2ko65') #globito
  for i in textfinder1:
    text += i.text + '\n'
  print(text)

def espera_barra_progreso():
   try:
    wait = WebDriverWait(driver, 40)
    wait.until(EC.invisibility_of_element_located((By.XPATH,'//*[@id="startup"]/div/progress')))
   except TimeoutException:
    print("esperamos.....")
    espera_barra_progreso()
  

def filtra_envia_contactos(contact, text, opc):
  contact = contact.replace("-", "")
  contact = contact.replace(" ","")
  JS_enlace = "window.location.href = 'https://web.whatsapp.com/send?phone=" + contact + "';"
  driver.execute_script(JS_enlace)
  time.sleep(4)
  espera_barra_progreso()
  while True:
    try:
      #aca busco el imput para enviar
      inp_xpath = '//div[@class="_3u328 copyable-text selectable-text"][@contenteditable="true"][@data-tab="1"]'
      input_box = WebDriverWait(driver, 10).until(
      lambda driver: driver.find_element_by_xpath(inp_xpath))
      time.sleep(2)
      if opc == 2:
        input_box.send_keys(text + Keys.ENTER)
        ws['D' + str(i)] = "enviado"
        print (contact + " si")
      else:
        ws['D' + str(i)] = "valido para enviar"
        print (contact + " si")
        pass
      time.sleep(2)
      file.save("contacts.xlsx")
      return
    except TimeoutException:
      print("no encontro el boton")
      no_encuentra = bool(driver.find_elements_by_xpath("//*[contains(text(), 'El número de teléfono compartido a través de la dirección URL es inválido')]"))
      if no_encuentra == True:
        ws['D' + str(i)] = "numero no valido"
        print (contact + " no")
        time.sleep(2)
        file.save("contacts.xlsx")
      return

def envia_respuesta(contact, text):
  inp_xpath_search = "//input[@title='Buscar o empezar un chat nuevo']"
  input_box_search = WebDriverWait(driver,50).until(lambda driver: driver.find_element_by_xpath(inp_xpath_search))
  input_box_search.click()
  time.sleep(2)
  input_box_search.send_keys(contact)
  time.sleep(5)
  while True:
    try:
      selected_contact = driver.find_element_by_xpath("//span[@title='"+'+'+contact+"']")
      selected_contact.click()
      inp_xpath = '//div[@class="_3u328 copyable-text selectable-text"][@contenteditable="true"][@data-tab="1"]'
      input_box = driver.find_element_by_xpath(inp_xpath)
      time.sleep(2)
      input_box.send_keys(text + Keys.ENTER)
      time.sleep(2)
      #bot_clear = '//button[@class="_2heX1"]'
      #time.sleep(5)
      #boton_clear = driver.find_element_by_xpath(bot_clear)
      #time.sleep(2)
      #boton_clear.click()
      ws['D' + str(i)] = "enviado"
      file.save("contacts.xlsx")
      time.sleep(2)
      return
    except NoSuchElementException:
      print("Oops! no esta ese contacto...")
      time.sleep(2)
      bot_clear = '//button[@class="_2heX1"]'
      time.sleep(5)
      boton_clear = driver.find_element_by_xpath(bot_clear)
      time.sleep(2)
      boton_clear.click()
      ws['D' + str(i)] = "no enviado"
      file.save("contacts.xlsx")
      return


driver = webdriver.Chrome('chromedriver.exe')
driver.get('https://web.whatsapp.com/')

while True:
  try:
    qr = driver.find_element_by_class_name("landing-main")
    input("Escanee el Qr")
  except:
    opcion = 0
    op = menu_principal(opcion)
    break
if op == 4:
  busca_elemetos()
  exit


i = 0
file = excel.load_workbook(filename = "contacts.xlsx")
ws = file.active
for row in ws.values:
    i = i + 1
    if "".__eq__(str(row[0])):
      pass
    else:
      con = str(row[0])
      if con[0] == "0":
        con = con.replace(con[0],"")
      contacto ="54 9 " + con
      texto = row[1] +", " + row[2]
      print("\n" + contacto + "\n" + texto)
      if op == 1:
        filtra_envia_contactos(contacto, texto, op)
      elif op == 2:
        filtra_envia_contactos(contacto, texto, op)
      elif op == 3:
        envia_respuesta(contacto, texto)
  
    